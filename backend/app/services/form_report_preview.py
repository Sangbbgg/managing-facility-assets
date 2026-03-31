from html import escape
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.form_template import ReportFormTemplate, ReportFormMapping
from app.services.form_report_builder import (
    _get_defined_name_map,
    _resolve_mapping_anchor,
    _resolve_page_settings,
    _resolve_subblock_layout,
    parse_cell,
    fetch_data_source,
    format_value,
    resolve_mapping_value,
    _resolve_repeat_target,
)


def _color_to_hex(color) -> str | None:
    if color is None:
        return None
    if hasattr(color, 'rgb') and color.rgb and color.rgb != "00000000":
        rgb = str(color.rgb)
        if len(rgb) == 8:
            return f"#{rgb[2:]}"
        return f"#{rgb}"
    return None


def _cell_style_css(cell) -> str:
    parts = []
    if cell.fill and cell.fill.fgColor:
        bg = _color_to_hex(cell.fill.fgColor)
        if bg:
            parts.append(f"background:{bg}")
    if cell.font:
        if cell.font.bold:
            parts.append("font-weight:500")
        if cell.font.size:
            parts.append(f"font-size:{max(cell.font.size * 0.85, 11)}px")
        fc = _color_to_hex(cell.font.color) if cell.font.color else None
        if fc:
            parts.append(f"color:{fc}")
    if cell.alignment:
        if cell.alignment.horizontal:
            parts.append(f"text-align:{cell.alignment.horizontal}")
        if cell.alignment.vertical:
            parts.append(f"vertical-align:{cell.alignment.vertical}")
    if cell.border:
        has_border = any([
            cell.border.left and cell.border.left.style,
            cell.border.right and cell.border.right.style,
            cell.border.top and cell.border.top.style,
            cell.border.bottom and cell.border.bottom.style,
        ])
        if has_border:
            parts.append("border:1px solid #C4A882")
    parts.append("padding:6px 8px")
    return ";".join(parts)


def _sheet_to_html(ws, cell_values: dict[tuple[int, int], str] | None = None) -> str:
    cell_values = cell_values or {}

    merged = {}
    for mr in ws.merged_cells.ranges:
        min_r, min_c = mr.min_row, mr.min_col
        max_r, max_c = mr.max_row, mr.max_col
        merged[(min_r, min_c)] = (max_r - min_r + 1, max_c - min_c + 1)
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if (r, c) != (min_r, min_c):
                    merged[(r, c)] = None

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    html_parts = [
        '<table style="border-collapse:collapse;table-layout:fixed;min-width:max-content;'
        'font-family:\'Malgun Gothic\',sans-serif;font-size:12px;background:#fff;">'
    ]

    for r in range(1, max_row + 1):
        html_parts.append("<tr>")
        for c in range(1, max_col + 1):
            merge_info = merged.get((r, c), "no_merge")
            if merge_info is None:
                continue
            cell = ws.cell(r, c)
            css = _cell_style_css(cell)
            attrs = f'style="{css};min-width:72px;max-width:240px;white-space:pre-wrap;word-break:break-word;"'
            if isinstance(merge_info, tuple):
                rowspan, colspan = merge_info
                if rowspan > 1:
                    attrs += f' rowspan="{rowspan}"'
                if colspan > 1:
                    attrs += f' colspan="{colspan}"'
            value = cell_values.get((r, c))
            if value is None:
                value = cell.value if cell.value is not None else ""
            html_parts.append(f"<td {attrs}>{escape(str(value))}</td>")
        html_parts.append("</tr>")

    html_parts.append("</table>")
    return "\n".join(html_parts)


async def generate_preview_html(
    template_id: int,
    asset_id: int,
    db: AsyncSession,
) -> str:
    template = await db.get(ReportFormTemplate, template_id)
    if not template:
        raise ValueError("양식을 찾을 수 없습니다")

    stmt = (
        select(ReportFormMapping)
        .where(ReportFormMapping.template_id == template_id)
        .order_by(ReportFormMapping.sort_order)
    )
    mappings = (await db.execute(stmt)).scalars().all()

    data_cache = {}

    async def get_data(ds):
        if ds not in data_cache:
            data_cache[ds] = await fetch_data_source(ds, asset_id, db)
        return data_cache[ds]

    wb = load_workbook(template.file_path)
    defined_names = _get_defined_name_map(template.file_path)
    ws = wb.active

    cell_values = {}

    for m in mappings:
        if m.repeat_direction:
            continue
        resolved_sheet_name, row, col = _resolve_mapping_anchor(m, defined_names)
        if resolved_sheet_name and resolved_sheet_name in wb.sheetnames:
            ws = wb[resolved_sheet_name]
        if m.data_source == "static":
            cell_values[(row, col)] = m.field
            continue
        data = await get_data(m.data_source)
        cell_values[(row, col)] = resolve_mapping_value(data, m)

    repeat_groups = {}
    for m in mappings:
        if not m.repeat_direction:
            continue
        resolved_sheet_name, row, col = _resolve_mapping_anchor(m, defined_names)
        direction = m.repeat_direction or "down"
        anchor = row if direction == "down" else col
        key = (resolved_sheet_name or m.sheet_name or "", m.data_source, direction, anchor)
        if key not in repeat_groups:
            repeat_groups[key] = []
        repeat_groups[key].append((m, row, col))

    for (_sheet_name, ds, direction, _anchor), group in repeat_groups.items():
        data = await get_data(ds)
        if not isinstance(data, list):
            continue
        max_rows = group[0][0].repeat_max_rows
        if max_rows is None or max_rows <= 0:
            max_rows = len(data)
        for i, item in enumerate(data[:max_rows]):
            for m, base_row, base_col in group:
                if _sheet_name and _sheet_name in wb.sheetnames:
                    ws = wb[_sheet_name]
                if direction == "down" and getattr(m, "overflow_mode", None) == "sheet_right":
                    page_settings = _resolve_page_settings(m, defined_names)
                    block_start_row = page_settings["block_start_row"]
                    block_end_row = page_settings["block_end_row"]
                    block_start_col = page_settings["block_start_col"]
                    block_end_col = page_settings["block_end_col"]
                    page_start_col = page_settings["page_start_col"]
                    page_end_col = page_settings["page_end_col"]
                    if all([block_start_row, block_end_row, block_start_col, block_end_col]):
                        rows_per_block = max(block_end_row - block_start_row + 1, 1)
                        start_col_idx = column_index_from_string(block_start_col)
                        end_col_idx = column_index_from_string(block_end_col)
                        body_width = max(end_col_idx - start_col_idx + 1, 1)
                        page_start_col_idx = column_index_from_string(page_start_col) if page_start_col else start_col_idx
                        page_end_col_idx = column_index_from_string(page_end_col) if page_end_col else end_col_idx
                        page_width = max(page_end_col_idx - page_start_col_idx + 1, 1)
                        subblock_count, subblock_width = _resolve_subblock_layout(m, body_width, page_width)
                        block_fill_mode = len(group) == 1 and body_width > 1
                        if block_fill_mode:
                            page_capacity = rows_per_block * body_width * subblock_count
                            page_index = i // page_capacity
                            page_item_index = i % page_capacity
                            subblock_index = page_item_index // (rows_per_block * body_width)
                            block_item_index = page_item_index % (rows_per_block * body_width)
                            row_offset = block_item_index // body_width
                            col_in_body = block_item_index % body_width
                            col_offset = (page_index * page_width) + (subblock_index * subblock_width) + col_in_body
                            target_col = start_col_idx + col_offset
                        else:
                            page_capacity = rows_per_block * subblock_count
                            page_index = i // page_capacity
                            page_item_index = i % page_capacity
                            subblock_index = page_item_index // rows_per_block
                            row_offset = page_item_index % rows_per_block
                            col_offset = (page_index * page_width) + (subblock_index * subblock_width)
                            target_col = base_col + col_offset
                        target_row = block_start_row + row_offset
                    else:
                        target_row, target_col = _resolve_repeat_target(ws, base_row, base_col, direction, i)
                else:
                    target_row, target_col = _resolve_repeat_target(ws, base_row, base_col, direction, i)
                cell_values[(target_row, target_col)] = resolve_mapping_value(
                    item,
                    m,
                    count_value=len(data),
                )

    return _sheet_to_html(ws, cell_values)


async def analyze_template_structure(file_path: str) -> dict:
    wb = load_workbook(file_path)
    ws = wb.active

    merged_cells = []
    for mr in ws.merged_cells.ranges:
        merged_cells.append({
            "range": str(mr),
            "start_cell": f"{get_column_letter(mr.min_col)}{mr.min_row}",
            "rows": mr.max_row - mr.min_row + 1,
            "cols": mr.max_col - mr.min_col + 1,
        })

    label_cells = []
    for r in range(1, (ws.max_row or 0) + 1):
        for c in range(1, (ws.max_column or 0) + 1):
            cell = ws.cell(r, c)
            if cell.value and isinstance(cell.value, str) and len(cell.value.strip()) > 0:
                label_cells.append({
                    "cell": f"{get_column_letter(c)}{r}",
                    "value": str(cell.value)[:100],
                    "has_fill": bool(cell.fill and cell.fill.fgColor and _color_to_hex(cell.fill.fgColor)),
                })

    defined_names = list(_get_defined_name_map(file_path).values())

    return {
        "sheet_name": ws.title,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "merged_cells": merged_cells,
        "label_cells": label_cells[:200],
        "defined_names": defined_names,
    }


async def generate_template_workbook_preview(file_path: str) -> dict:
    wb = load_workbook(file_path)
    sheets = []

    for ws in wb.worksheets:
        sheets.append({
            "name": ws.title,
            "max_row": ws.max_row or 1,
            "max_col": ws.max_column or 1,
            "html": _sheet_to_html(ws),
        })

    return {
        "active_sheet": wb.active.title if wb.worksheets else None,
        "sheet_count": len(sheets),
        "sheets": sheets,
    }
