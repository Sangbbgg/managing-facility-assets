import io
import re
import tempfile
from fastapi import UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.services.asset_code import issue_asset_code
from app.models.master import GroupNode, EquipmentType, LocationNode
from app.models.asset import Asset, AssetCodeSequence
from app.core.config import settings


IMPORTANCE_VALUES = ["상", "중", "하"]

TEMPLATE_HEADERS = [
    "자산명*", "그룹*", "장비종류*", "위치", "중요도", "설치일(YYYY-MM-DD)", "자산코드(자동)"
]

REF_SHEET = "자산코드 참고"


def _asset_code_formula(row: int, ref_last_row: int) -> str:
    """
    자산코드 자동 계산 수식 (Excel).

    - B열: "full_path (GROUP_CODE)" 형식에서 코드 추출
    - C열: "name (TYPE_CODE)" 형식에서 코드 추출
    - '자산코드 참고' 시트의 E열(현재 마지막 번호)을 SUMPRODUCT로 조회
    - 같은 그룹+장비종류 조합이 이 행 위에 몇 번 나왔는지 COUNTIFS로 누적
    """
    r = row
    N = ref_last_row
    P = settings.ASSET_PREFIX
    REF = REF_SHEET  # "자산코드 참고"

    # 그룹코드: MID(B{r}, FIND("(",B{r})+1, FIND(")",B{r})-FIND("(",B{r})-1)
    grp = f'MID(B{r},FIND("(",B{r})+1,FIND(")",B{r})-FIND("(",B{r})-1)'
    # 장비종류코드: MID(C{r}, ...)
    typ = f'MID(C{r},FIND("(",C{r})+1,FIND(")",C{r})-FIND("(",C{r})-1)'

    # '자산코드 참고' 시트 참조 (A열=그룹코드, C열=장비종류코드, E열=last_seq)
    ref_grp = f"'{REF}'!$A$2:$A${N}"
    ref_typ = f"'{REF}'!$C$2:$C${N}"
    ref_seq = f"'{REF}'!$E$2:$E${N}"

    # SUMPRODUCT로 last_seq 조회
    lookup = f"SUMPRODUCT(({ref_grp}={grp})*({ref_typ}={typ})*{ref_seq})"

    # 이 행 위에 같은 조합이 몇 번 나왔는지 (row 2는 0)
    if row == 2:
        above_count = "0"
    else:
        above_count = f'COUNTIFS($B$2:B{r-1},B{r},$C$2:C{r-1},C{r})'

    seq = f"{lookup}+1+{above_count}"

    return (
        f'=IF(OR(B{r}="",C{r}=""),"",'
        f'"{P}-"&{grp}&"-"&{typ}&"-"&TEXT({seq},"0000"))'
    )


async def generate_template(db: AsyncSession) -> str:
    # ── 데이터 조회 ──────────────────────────────────────────────────────────
    groups_result = await db.execute(
        select(GroupNode).where(GroupNode.code.isnot(None)).order_by(GroupNode.id)
    )
    groups = groups_result.scalars().all()

    types_result = await db.execute(select(EquipmentType).order_by(EquipmentType.id))
    eq_types = types_result.scalars().all()

    locations_result = await db.execute(select(LocationNode).order_by(LocationNode.depth, LocationNode.id))
    all_locations = locations_result.scalars().all()
    parent_ids = {loc.parent_id for loc in all_locations if loc.parent_id is not None}
    leaf_locations = [loc for loc in all_locations if loc.id not in parent_ids]

    seq_result = await db.execute(select(AssetCodeSequence))
    seq_rows = {(s.group_code, s.type_code): s.last_seq for s in seq_result.scalars().all()}

    ref_data_rows = len(groups) * len(eq_types)  # 자산코드 참고 시트 데이터 행 수
    ref_last_row = ref_data_rows + 1              # 헤더 1행 + 데이터

    wb = Workbook()

    # ── 1. 자산등록 시트 ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "자산등록"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    auto_fill   = PatternFill("solid", fgColor="E2EFDA")  # 연두: 자동 계산 열
    auto_font   = Font(color="375623", bold=True)

    for col, header in enumerate(TEMPLATE_HEADERS, 1):
        cell = ws.cell(1, col, header)
        if col == 7:
            cell.fill = auto_fill
            cell.font = auto_font
        else:
            cell.fill = header_fill
            cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 42
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 28

    # 자산코드 수식 (행 2~201, 최대 200건)
    MAX_DATA_ROWS = 200
    for r in range(2, MAX_DATA_ROWS + 2):
        cell = ws.cell(r, 7, _asset_code_formula(r, ref_last_row))
        cell.fill = PatternFill("solid", fgColor="F0F7EE")
        cell.alignment = Alignment(horizontal="center")

    # ── 2. 목록 시트 (숨김) ────────────────────────────────────────────────
    ws_list = wb.create_sheet("목록")

    group_labels = [f"{g.full_path} ({g.code})" for g in groups]
    for i, label in enumerate(group_labels, 1):
        ws_list.cell(i, 1, label)

    type_labels = [f"{t.name} ({t.code})" for t in eq_types]
    for i, label in enumerate(type_labels, 1):
        ws_list.cell(i, 2, label)

    loc_labels = [loc.full_path or loc.name for loc in leaf_locations]
    for i, label in enumerate(loc_labels, 1):
        ws_list.cell(i, 3, label)

    for i, val in enumerate(IMPORTANCE_VALUES, 1):
        ws_list.cell(i, 4, val)

    ws_list.sheet_state = "hidden"

    # ── 드롭다운 DataValidation ──────────────────────────────────────────────
    def make_dv(formula: str, col: int):
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        col_letter = get_column_letter(col)
        dv.sqref = f"{col_letter}2:{col_letter}{MAX_DATA_ROWS + 1}"

    if group_labels:
        make_dv(f"목록!$A$1:$A${len(group_labels)}", 2)
    if type_labels:
        make_dv(f"목록!$B$1:$B${len(type_labels)}", 3)
    if loc_labels:
        make_dv(f"목록!$C$1:$C${len(loc_labels)}", 4)
    make_dv(f"목록!$D$1:$D${len(IMPORTANCE_VALUES)}", 5)

    # 예시 행
    ws.cell(2, 1, "예시_서버01")
    if group_labels:
        ws.cell(2, 2, group_labels[0])
    if type_labels:
        ws.cell(2, 3, type_labels[0])
    if loc_labels:
        ws.cell(2, 4, loc_labels[0])
    ws.cell(2, 5, "중")

    # ── 3. 자산코드 참고 시트 ──────────────────────────────────────────────
    ws_ref = wb.create_sheet(REF_SHEET)
    ref_headers = ["그룹코드", "그룹명(full_path)", "장비종류코드", "장비종류명", "현재 마지막 번호", "다음 자산코드"]
    hfill = PatternFill("solid", fgColor="70AD47")
    for col, h in enumerate(ref_headers, 1):
        cell = ws_ref.cell(1, col, h)
        cell.fill = hfill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    ws_ref.column_dimensions["A"].width = 12
    ws_ref.column_dimensions["B"].width = 40
    ws_ref.column_dimensions["C"].width = 14
    ws_ref.column_dimensions["D"].width = 18
    ws_ref.column_dimensions["E"].width = 20
    ws_ref.column_dimensions["F"].width = 30

    row = 2
    for g in groups:
        for t in eq_types:
            last_seq = seq_rows.get((g.code, t.code), 0)
            next_seq = last_seq + 1
            next_code = f"{settings.ASSET_PREFIX}-{g.code}-{t.code}-{next_seq:04d}"
            ws_ref.cell(row, 1, g.code)
            ws_ref.cell(row, 2, g.full_path)
            ws_ref.cell(row, 3, t.code)
            ws_ref.cell(row, 4, t.name)
            ws_ref.cell(row, 5, last_seq)
            ws_ref.cell(row, 6, next_code)
            row += 1

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return tmp.name


# ── 파싱 & 등록 ──────────────────────────────────────────────────────────────

_CODE_RE = re.compile(r"\((\w+)\)\s*$")


def _extract_code(value: str) -> str | None:
    if not value:
        return None
    m = _CODE_RE.search(str(value).strip())
    return m.group(1) if m else str(value).strip()


async def parse_and_register(file: UploadFile, db: AsyncSession) -> dict:
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    loc_result = await db.execute(select(LocationNode))
    location_map = {loc.full_path: loc.id for loc in loc_result.scalars().all() if loc.full_path}

    results = {"success": 0, "errors": []}

    for row_idx in range(2, ws.max_row + 1):
        try:
            asset_name   = ws.cell(row_idx, 1).value
            group_val    = ws.cell(row_idx, 2).value
            type_val     = ws.cell(row_idx, 3).value
            location_val = ws.cell(row_idx, 4).value
            importance   = ws.cell(row_idx, 5).value or "중"
            install_date = ws.cell(row_idx, 6).value
            # G열(자산코드 자동)은 무시 — 서버에서 새로 채번

            if not asset_name and not group_val and not type_val:
                break

            if not asset_name or not group_val or not type_val:
                results["errors"].append({"row": row_idx, "error": "자산명·그룹·장비종류는 필수입니다"})
                continue

            group_code = _extract_code(str(group_val))
            type_code  = _extract_code(str(type_val))

            group = await db.scalar(select(GroupNode).where(GroupNode.code == group_code))
            etype = await db.scalar(select(EquipmentType).where(EquipmentType.code == type_code))
            if not group:
                results["errors"].append({"row": row_idx, "error": f"그룹 '{group_val}' 없음"})
                continue
            if not etype:
                results["errors"].append({"row": row_idx, "error": f"장비종류 '{type_val}' 없음"})
                continue

            location_id = None
            if location_val:
                loc_fp = str(location_val).strip()
                location_id = location_map.get(loc_fp)
                if location_id is None:
                    results["errors"].append({"row": row_idx, "error": f"위치 '{loc_fp}' 없음"})
                    continue

            asset_code = await issue_asset_code(db, group_code, type_code)

            asset = Asset(
                asset_code=asset_code,
                asset_name=str(asset_name),
                group_id=group.id,
                equipment_type_id=etype.id,
                location_id=location_id,
                importance=str(importance) if importance else "중",
                install_date=install_date if hasattr(install_date, 'year') else None,
            )
            db.add(asset)
            await db.flush()
            results["success"] += 1

        except Exception as e:
            results["errors"].append({"row": row_idx, "error": str(e)})

    return results
