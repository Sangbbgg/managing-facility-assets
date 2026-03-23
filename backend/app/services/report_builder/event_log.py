import tempfile
from datetime import date
from collections import defaultdict
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from app.models.record import EventLogRecord
from app.models.asset import Asset
from ._common import write_header, style_data_row

HEADERS = ["자산코드", "설비명", "날짜", "로그유형", "이벤트ID", "레벨", "건수"]
WIDTHS  = [18, 20, 12, 10, 12, 10, 10]
LEVEL_MAP = {1: "Critical", 2: "Error", 3: "Warning"}


async def build(date_from: date, date_to: date, db: AsyncSession) -> str:
    stmt = (
        select(EventLogRecord, Asset)
        .join(Asset, EventLogRecord.asset_id == Asset.id)
        .where(
            EventLogRecord.record_date >= date_from,
            EventLogRecord.record_date <= date_to,
        )
        .order_by(EventLogRecord.record_date, Asset.asset_code)
    )
    rows = (await db.execute(stmt)).all()

    monthly: dict[str, list] = defaultdict(list)
    for rec, asset in rows:
        monthly[rec.record_date.strftime("%Y년 %m월")].append((rec, asset))

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "전체요약"
    ws_sum.cell(1, 1, f"조회 기간: {date_from} ~ {date_to}")
    ws_sum.cell(2, 1, f"총 이벤트 유형 수: {len(rows)}")

    for month_key, items in sorted(monthly.items()):
        ws = wb.create_sheet(title=month_key)
        write_header(ws, HEADERS, WIDTHS)
        for i, (rec, asset) in enumerate(items, 2):
            ws.cell(i, 1, asset.asset_code)
            ws.cell(i, 2, asset.asset_name)
            ws.cell(i, 3, str(rec.record_date))
            ws.cell(i, 4, rec.log_type)
            ws.cell(i, 5, rec.event_id)
            ws.cell(i, 6, LEVEL_MAP.get(rec.level, str(rec.level)))
            ws.cell(i, 7, rec.count)
            style_data_row(ws, i, len(HEADERS))

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return tmp.name
