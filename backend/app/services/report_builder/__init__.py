import tempfile
from typing import Optional
from collections import defaultdict
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from app.models.report import ReportTemplate
from .field_catalog import MULTI_SHEET_SOURCES
from .data_fetchers import fetch_rows


def _border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


BORDER   = _border()
DATA_FONT = Font(size=10)


def _write_sheet(ws, template: ReportTemplate, rows: list[dict]):
    columns = template.columns  # [{header, field, width}]
    fill_hex = template.header_color.lstrip("#") if template.header_color else "1F4E79"
    header_fill = PatternFill(fill_type="solid", fgColor=fill_hex)
    header_font = Font(color="FFFFFF", bold=True, size=10)
    center      = Alignment(horizontal="center", vertical="center")

    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(1, col_idx, col["header"])
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center
        cell.border    = BORDER
        ws.column_dimensions[cell.column_letter].width = col.get("width", 15)

    for row_idx, row in enumerate(rows, 2):
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row_idx, col_idx, row.get(col["field"], ""))
            cell.font      = DATA_FONT
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="center")


async def build_report(template_id: int, year: int, month: Optional[int], db: AsyncSession) -> str:
    template = await db.get(ReportTemplate, template_id)
    if not template:
        raise ValueError(f"보고서 서식을 찾을 수 없습니다 (id={template_id})")

    period_label = f"{year}년{f' {month}월' if month else ' 전체'}"
    rows = await fetch_rows(template.data_source, year, month, db)

    wb = Workbook()

    if template.data_source in MULTI_SHEET_SOURCES and rows and "_sheet_key" in rows[0]:
        # 월별/분기별 시트 분리
        ws_sum = wb.active
        ws_sum.title = "전체요약"
        ws_sum.cell(1, 1, f"조회 기간: {period_label}")
        ws_sum.cell(2, 1, f"총 건수: {len(rows)}")

        grouped: dict[str, list] = defaultdict(list)
        for row in rows:
            grouped[row["_sheet_key"]].append(row)

        for sheet_key in sorted(grouped):
            ws = wb.create_sheet(title=sheet_key)
            _write_sheet(ws, template, grouped[sheet_key])
    else:
        ws = wb.active
        ws.title = template.sheet_name or "Sheet1"
        _write_sheet(ws, template, rows)

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return tmp.name
