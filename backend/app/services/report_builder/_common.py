from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
DATA_FONT = Font(size=10)
CENTER = Alignment(horizontal="center", vertical="center")


def write_header(ws, headers: list[str], widths: list[int]):
    for col, text in enumerate(headers, 1):
        cell = ws.cell(1, col, text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w


def style_data_row(ws, row: int, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row, col)
        cell.font = DATA_FONT
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center")
