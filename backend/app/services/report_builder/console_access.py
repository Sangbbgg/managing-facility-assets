import tempfile
from datetime import date
from collections import defaultdict
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from app.models.record import ConsoleAccessRecord
from app.models.asset import Asset
from ._common import write_header, style_data_row

HEADERS = ["자산코드", "설비명", "접속일", "접속자", "목적"]
WIDTHS  = [18, 20, 12, 12, 30]


async def build(date_from: date, date_to: date, db: AsyncSession) -> str:
    stmt = (
        select(ConsoleAccessRecord, Asset)
        .join(Asset, ConsoleAccessRecord.asset_id == Asset.id)
        .where(
            ConsoleAccessRecord.access_date >= date_from,
            ConsoleAccessRecord.access_date <= date_to,
        )
        .order_by(ConsoleAccessRecord.access_date)
    )
    rows = (await db.execute(stmt)).all()

    monthly: dict[str, list] = defaultdict(list)
    for rec, asset in rows:
        monthly[rec.access_date.strftime("%Y년 %m월")].append((rec, asset))

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "전체요약"
    ws_sum.cell(1, 1, f"조회 기간: {date_from} ~ {date_to}")
    ws_sum.cell(2, 1, f"총 접속 건수: {len(rows)}")

    for month_key, items in sorted(monthly.items()):
        ws = wb.create_sheet(title=month_key)
        write_header(ws, HEADERS, WIDTHS)
        for i, (rec, asset) in enumerate(items, 2):
            ws.cell(i, 1, asset.asset_code)
            ws.cell(i, 2, asset.asset_name)
            ws.cell(i, 3, str(rec.access_date))
            ws.cell(i, 4, rec.accessor)
            ws.cell(i, 5, rec.purpose or "")
            style_data_row(ws, i, len(HEADERS))

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return tmp.name
