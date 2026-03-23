import tempfile
from datetime import date
from collections import defaultdict
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from app.models.record import InspectionRecord
from app.models.asset import Asset
from ._common import write_header, style_data_row

HEADERS = ["자산코드", "설비명", "점검일", "점검유형", "결과", "점검자", "특이사항"]
WIDTHS  = [18, 20, 12, 10, 8, 12, 30]


async def build(date_from: date, date_to: date, db: AsyncSession) -> str:
    stmt = (
        select(InspectionRecord, Asset)
        .join(Asset, InspectionRecord.asset_id == Asset.id)
        .where(
            InspectionRecord.record_date >= date_from,
            InspectionRecord.record_date <= date_to,
            InspectionRecord.inspection_type == "MONTHLY",
        )
        .order_by(InspectionRecord.record_date)
    )
    rows = (await db.execute(stmt)).all()

    monthly: dict[str, list] = defaultdict(list)
    for record, asset in rows:
        monthly[record.record_date.strftime("%Y년 %m월")].append((record, asset))

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "전체요약"
    ws_sum.cell(1, 1, f"조회 기간: {date_from} ~ {date_to}")
    ws_sum.cell(2, 1, f"총 점검 건수: {len(rows)}")

    for month_key, items in sorted(monthly.items()):
        ws = wb.create_sheet(title=month_key)
        write_header(ws, HEADERS, WIDTHS)
        for i, (rec, asset) in enumerate(items, 2):
            ws.cell(i, 1, asset.asset_code)
            ws.cell(i, 2, asset.asset_name)
            ws.cell(i, 3, str(rec.record_date))
            ws.cell(i, 4, rec.inspection_type)
            ws.cell(i, 5, rec.result or "")
            ws.cell(i, 6, rec.inspector or "")
            ws.cell(i, 7, rec.special_notes or "")
            style_data_row(ws, i, len(HEADERS))

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return tmp.name
