import tempfile
from datetime import date
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from app.models.asset import Asset
from app.models.master import Person, Department
from ._common import write_header, style_data_row

HEADERS = ["자산코드", "설비명", "담당자", "직책", "부서", "연락처", "책임자", "상태"]
WIDTHS  = [18, 20, 12, 10, 16, 18, 12, 10]


async def build(date_from: date, date_to: date, db: AsyncSession) -> str:
    stmt = select(Asset).where(Asset.is_deleted == False).order_by(Asset.asset_code)
    assets = (await db.execute(stmt)).scalars().all()

    persons = {r.id: r for r in (await db.execute(select(Person))).scalars()}
    depts   = {r.id: r for r in (await db.execute(select(Department))).scalars()}

    wb = Workbook()
    ws = wb.active
    ws.title = "설비관리대장"
    write_header(ws, HEADERS, WIDTHS)

    for i, a in enumerate(assets, 2):
        mgr = persons.get(a.manager_id)
        sup = persons.get(a.supervisor_id)
        dept = depts.get(mgr.dept_id) if mgr and mgr.dept_id else None
        ws.cell(i, 1, a.asset_code)
        ws.cell(i, 2, a.asset_name)
        ws.cell(i, 3, mgr.name if mgr else "")
        ws.cell(i, 4, mgr.title if mgr else "")
        ws.cell(i, 5, dept.name if dept else "")
        ws.cell(i, 6, mgr.contact if mgr else "")
        ws.cell(i, 7, sup.name if sup else "")
        ws.cell(i, 8, a.status)
        style_data_row(ws, i, len(HEADERS))

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return tmp.name
