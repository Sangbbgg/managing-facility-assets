import tempfile
from datetime import date
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from app.models.record import PasswordRecord
from app.models.asset import Asset
from ._common import write_header, style_data_row

HEADERS = ["자산코드", "설비명", "계정명", "변경일", "변경자", "비고"]
WIDTHS  = [18, 20, 16, 12, 12, 25]


async def build(date_from: date, date_to: date, db: AsyncSession) -> str:
    stmt = (
        select(PasswordRecord, Asset)
        .join(Asset, PasswordRecord.asset_id == Asset.id)
        .where(
            PasswordRecord.changed_date >= date_from,
            PasswordRecord.changed_date <= date_to,
        )
        .order_by(PasswordRecord.changed_date)
    )
    rows = (await db.execute(stmt)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "비밀번호관리대장"
    write_header(ws, HEADERS, WIDTHS)

    for i, (rec, asset) in enumerate(rows, 2):
        ws.cell(i, 1, asset.asset_code)
        ws.cell(i, 2, asset.asset_name)
        ws.cell(i, 3, rec.account_name)
        ws.cell(i, 4, str(rec.changed_date))
        ws.cell(i, 5, rec.changed_by or "")
        ws.cell(i, 6, rec.notes or "")
        style_data_row(ws, i, len(HEADERS))

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return tmp.name
