import tempfile
from datetime import date

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.asset import Asset
from app.models.hw_info import AssetHwNic
from app.models.master import EquipmentType, GroupNode, LocationNode
from ._common import style_data_row, write_header

HEADERS = ["자산코드", "설비명", "그룹", "위치", "장비종류", "IP", "중요도", "설치일", "상태"]
WIDTHS = [18, 20, 20, 24, 14, 16, 8, 12, 10]


async def build(date_from: date, date_to: date, db: AsyncSession) -> str:
    stmt = select(Asset).where(Asset.is_deleted == False).order_by(Asset.asset_code)
    assets = (await db.execute(stmt)).scalars().all()

    groups = {row.id: row for row in (await db.execute(select(GroupNode))).scalars()}
    locations = {row.id: row for row in (await db.execute(select(LocationNode))).scalars()}
    equipment_types = {row.id: row for row in (await db.execute(select(EquipmentType))).scalars()}
    nics = {row.id: row for row in (await db.execute(select(AssetHwNic))).scalars()}

    wb = Workbook()
    ws = wb.active
    ws.title = "자산현황"
    write_header(ws, HEADERS, WIDTHS)

    for index, asset in enumerate(assets, 2):
        ws.cell(index, 1, asset.asset_code)
        ws.cell(index, 2, asset.asset_name)
        ws.cell(index, 3, groups[asset.group_id].name if asset.group_id and asset.group_id in groups else "")
        ws.cell(index, 4, locations[asset.location_id].full_path if asset.location_id and asset.location_id in locations else "")
        ws.cell(index, 5, equipment_types[asset.equipment_type_id].name if asset.equipment_type_id and asset.equipment_type_id in equipment_types else "")
        nic = nics.get(asset.representative_nic_id) if asset.representative_nic_id else None
        ws.cell(index, 6, nic.ipv4_address if nic else "")
        ws.cell(index, 7, asset.importance or "")
        ws.cell(index, 8, str(asset.install_date) if asset.install_date else "")
        ws.cell(index, 9, asset.status)
        style_data_row(ws, index, len(HEADERS))

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return tmp.name
