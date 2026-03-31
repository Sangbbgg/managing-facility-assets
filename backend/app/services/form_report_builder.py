import re
import tempfile
import zipfile
import ast
import operator
from copy import copy
from datetime import date, datetime
from typing import Any
import json
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.asset import Asset
from app.models.collection import AssetCollectRun, AssetNetworkConnection
from app.models.form_template import ReportFormMapping, ReportFormTemplate
from app.models.hw_info import AssetHwCpu, AssetHwDisk, AssetHwGpu, AssetHwMemory, AssetHwNic, AssetHwOptical, AssetHwSystem
from app.models.record import ConsoleAccessRecord, EventLogRecord, InspectionRecord, PasswordRecord, SealRecord
from app.models.sw_info import AssetSwAccount, AssetSwHotfix, AssetSwProcess, AssetSwProduct
from app.schemas.form_template import FormFieldInfo


_EXCLUDED_FIELDS = {
    "id",
    "asset_id",
    "parent_id",
    "group_id",
    "location_id",
    "equipment_type_id",
    "manager_id",
    "dept_id",
    "representative_nic_id",
    "representative_account_id",
    "is_deleted",
    "depth",
    "code",
    "key",
    "raw_json",
    "raw_payload",
    "raw_csproduct",
    "raw_os",
    "raw_systeminfo",
}

_ASSET_DERIVED_FIELDS = [
    ("group_name", "Group Name"),
    ("group_full_path", "Group Full Path"),
    ("location_name", "Location Name"),
    ("location_full_path", "Location Full Path"),
    ("equipment_type_name", "Equipment Type Name"),
    ("manager_name", "Manager Name"),
    ("manager_title", "Manager Title"),
    ("manager_contact", "Manager Contact"),
    ("ip_address", "Representative IP"),
]

_DATA_SOURCE_MODELS: dict[str, tuple[Any, bool, Any | None]] = {
    "hw_system": (AssetHwSystem, False, None),
    "hw_cpu": (AssetHwCpu, True, AssetHwCpu.collected_at.desc()),
    "hw_memory": (AssetHwMemory, True, AssetHwMemory.collected_at.desc()),
    "hw_disk": (AssetHwDisk, True, AssetHwDisk.collected_at.desc()),
    "hw_optical": (AssetHwOptical, True, AssetHwOptical.collected_at.desc()),
    "hw_gpu": (AssetHwGpu, True, AssetHwGpu.collected_at.desc()),
    "hw_nic": (AssetHwNic, True, AssetHwNic.collected_at.desc()),
    "sw_product": (AssetSwProduct, True, AssetSwProduct.collected_at.desc()),
    "sw_hotfix": (AssetSwHotfix, True, AssetSwHotfix.installed_on.desc()),
    "sw_process": (AssetSwProcess, True, AssetSwProcess.collected_at.desc()),
    "sw_account": (AssetSwAccount, True, AssetSwAccount.collected_at.desc()),
    "network_connection": (AssetNetworkConnection, True, AssetNetworkConnection.collected_at.desc()),
    "collect_run": (AssetCollectRun, True, AssetCollectRun.collected_at.desc()),
    "inspection": (InspectionRecord, True, InspectionRecord.record_date.desc()),
    "event_log": (EventLogRecord, True, EventLogRecord.record_date.desc()),
    "console_access": (ConsoleAccessRecord, True, ConsoleAccessRecord.access_date.desc()),
    "seal": (SealRecord, True, SealRecord.record_date.desc()),
    "password": (PasswordRecord, True, PasswordRecord.changed_date.desc()),
}


def _humanize_label(field: str) -> str:
    return field.replace("_", " ").title()


def _flatten_custom_field_values(custom_fields: Any) -> dict[str, Any]:
    if not isinstance(custom_fields, dict):
        return {}
    flattened: dict[str, Any] = {}
    for key, value in custom_fields.items():
        key_text = str(key).strip()
        if not key_text:
            continue
        flattened[f"custom_fields.{key_text}"] = value
    return flattened


def _build_field_catalog() -> list[FormFieldInfo]:
    catalog: list[FormFieldInfo] = []

    for column in Asset.__table__.columns:
        if column.name in _EXCLUDED_FIELDS:
            continue
        catalog.append(
            FormFieldInfo(
                data_source="asset",
                field=column.name,
                label=_humanize_label(column.name),
            )
        )

    for field_name, label in _ASSET_DERIVED_FIELDS:
        catalog.append(
            FormFieldInfo(
                data_source="asset",
                field=field_name,
                label=label,
            )
        )

    for data_source, (model, is_repeatable, _order_by) in _DATA_SOURCE_MODELS.items():
        for column in model.__table__.columns:
            if column.name in _EXCLUDED_FIELDS:
                continue
            catalog.append(
                FormFieldInfo(
                    data_source=data_source,
                    field=column.name,
                    label=_humanize_label(column.name),
                    is_repeatable=is_repeatable,
                )
            )

    catalog.extend([
        FormFieldInfo(data_source="static", field="(direct-input)", label="Static Text", example="No issue found"),
        FormFieldInfo(data_source="today", field="date", label="Today Date", example="2026-03-30"),
        FormFieldInfo(data_source="today", field="year", label="Year", example="2026"),
        FormFieldInfo(data_source="today", field="month", label="Month", example="03"),
        FormFieldInfo(data_source="today", field="day", label="Day", example="30"),
        FormFieldInfo(data_source="today", field="weekday", label="Weekday", example="Monday"),
    ])

    return catalog


FIELD_CATALOG: list[FormFieldInfo] = _build_field_catalog()


def build_field_catalog(custom_field_keys: list[str] | None = None) -> list[FormFieldInfo]:
    catalog = list(FIELD_CATALOG)
    for key in custom_field_keys or []:
        key_text = str(key).strip()
        if not key_text:
            continue
        catalog.append(
            FormFieldInfo(
                data_source="asset",
                field=f"custom_fields.{key_text}",
                label=f"Custom Field · {key_text}",
            )
        )
    return catalog


def _stringify_preview_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return str(value)
    if isinstance(value, (list, dict)):
        try:
            return json.dumps(value, ensure_ascii=False)
        except TypeError:
            return str(value)
    return str(value)


def parse_cell(cell_str: str) -> tuple[int, int]:
    matched = re.match(r"([A-Z]+)(\d+)", cell_str.upper())
    if not matched:
        raise ValueError(f"invalid cell address: {cell_str}")
    col = column_index_from_string(matched.group(1))
    row = int(matched.group(2))
    return row, col


def _parse_defined_name_reference(reference: str) -> tuple[str, str] | None:
    ref = (reference or "").strip()
    if not ref:
        return None
    if ref.startswith("="):
        ref = ref[1:]
    if "!" not in ref:
        return None
    sheet_name, coord = ref.split("!", 1)
    sheet_name = sheet_name.strip().strip("'").replace("''", "'")
    coord = coord.strip().replace("$", "")
    if not coord:
        return None
    return sheet_name, coord


def _get_defined_name_map(file_path: str | None) -> dict[str, dict]:
    if not file_path:
        return {}

    workbook_ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    result: dict[str, dict] = {}

    with zipfile.ZipFile(file_path, "r") as archive:
        workbook_xml = archive.read("xl/workbook.xml")
    root = ET.fromstring(workbook_xml)

    sheet_names: list[str] = []
    for sheet_node in root.findall("main:sheets/main:sheet", workbook_ns):
        name = sheet_node.attrib.get("name")
        if name:
            sheet_names.append(name)

    for defined_node in root.findall("main:definedNames/main:definedName", workbook_ns):
        name = defined_node.attrib.get("name")
        if not name:
            continue
        parsed = _parse_defined_name_reference(defined_node.text or "")
        if not parsed:
            continue
        sheet_name, range_text = parsed
        local_sheet_id = defined_node.attrib.get("localSheetId")
        if local_sheet_id is not None and local_sheet_id.isdigit():
            index = int(local_sheet_id)
            if 0 <= index < len(sheet_names):
                sheet_name = sheet_names[index]
        if ":" in range_text:
            start_cell, end_cell = range_text.split(":", 1)
        else:
            start_cell = range_text
            end_cell = range_text
        start_col, start_row = coordinate_from_string(start_cell)
        end_col, end_row = coordinate_from_string(end_cell)
        result[name] = {
            "name": name,
            "sheet_name": sheet_name,
            "range": range_text,
            "start_cell": start_cell,
            "end_cell": end_cell,
            "start_row": start_row,
            "end_row": end_row,
            "start_col": start_col,
            "end_col": end_col,
            "rows": max(end_row - start_row + 1, 1),
            "cols": max(column_index_from_string(end_col) - column_index_from_string(start_col) + 1, 1),
            "local_sheet_id": int(local_sheet_id) if local_sheet_id is not None and local_sheet_id.isdigit() else None,
            "hidden": defined_node.attrib.get("hidden") == "1",
        }
    return result


def _resolve_mapping_anchor(mapping, defined_names: dict[str, dict]) -> tuple[str | None, int, int]:
    row, col = parse_cell(mapping.cell)
    named_range = getattr(mapping, "named_range_name", None)
    if named_range and named_range in defined_names:
        info = defined_names[named_range]
        return info["sheet_name"], row, col
    return mapping.sheet_name, row, col


def _resolve_page_settings(mapping, defined_names: dict[str, dict]) -> dict[str, str | int | None]:
    result = {
        "block_start_row": getattr(mapping, "block_start_row", None),
        "block_end_row": getattr(mapping, "block_end_row", None),
        "block_start_col": getattr(mapping, "block_start_col", None),
        "block_end_col": getattr(mapping, "block_end_col", None),
        "page_start_col": getattr(mapping, "block_start_col", None),
        "page_end_col": getattr(mapping, "block_end_col", None),
    }
    named_range = getattr(mapping, "named_range_name", None)
    if named_range and named_range in defined_names:
        body_info = defined_names[named_range]
        result["block_start_row"] = body_info["start_row"]
        result["block_end_row"] = body_info["end_row"]
        result["block_start_col"] = body_info["start_col"]
        result["block_end_col"] = body_info["end_col"]
        result["page_start_col"] = body_info["start_col"]
        result["page_end_col"] = body_info["end_col"]
    page_range_name = getattr(mapping, "page_range_name", None)
    if page_range_name and page_range_name in defined_names:
        page_info = defined_names[page_range_name]
        result["page_start_col"] = page_info["start_col"]
        result["page_end_col"] = page_info["end_col"]
    return result


def _is_merged_child(ws, row: int, col: int) -> bool:
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return not (mr.min_row == row and mr.min_col == col)
    return False


def _resolve_repeat_target(ws, base_row: int, base_col: int, direction: str, index: int) -> tuple[int, int]:
    if direction == "down":
        target_row = base_row + index
        while _is_merged_child(ws, target_row, base_col):
            target_row += 1
        return target_row, base_col

    target_col = base_col
    steps = index
    while steps > 0:
        target_col += 1
        while _is_merged_child(ws, base_row, target_col):
            target_col += 1
        steps -= 1
    return base_row, target_col


def _resolve_subblock_layout(mapping, body_width: int, page_width: int) -> tuple[int, int]:
    configured_count = getattr(mapping, "page_subblock_count", None)
    configured_width = getattr(mapping, "page_subblock_width", None)
    auto_count = max(page_width // body_width, 1) if body_width > 0 and page_width >= body_width and page_width % body_width == 0 else 1

    subblock_count = max(configured_count or auto_count, 1)
    if subblock_count == 1 and auto_count > 1:
        subblock_count = auto_count

    subblock_width = max(configured_width or body_width, 1)
    if (configured_width or 1) == 1 and body_width > 1:
        subblock_width = body_width

    return subblock_count, subblock_width


def _copy_cell_range(src_ws, tgt_ws, src_min_row: int, src_max_row: int, src_min_col: int, src_max_col: int, col_offset: int = 0) -> None:
    merge_targets: list[tuple[int, int, int, int]] = []
    for merge_range in list(src_ws.merged_cells.ranges):
        if (
            merge_range.min_row >= src_min_row
            and merge_range.max_row <= src_max_row
            and merge_range.min_col >= src_min_col
            and merge_range.max_col <= src_max_col
        ):
            merge_targets.append((
                merge_range.min_row,
                merge_range.min_col + col_offset,
                merge_range.max_row,
                merge_range.max_col + col_offset,
            ))

    for row in range(src_min_row, src_max_row + 1):
        target_row = row
        tgt_ws.row_dimensions[target_row].height = src_ws.row_dimensions[row].height
        for col in range(src_min_col, src_max_col + 1):
            if _is_merged_child(src_ws, row, col):
                continue
            if _is_merged_child(tgt_ws, row, col + col_offset):
                continue
            src_cell = src_ws.cell(row, col)
            tgt_cell = tgt_ws.cell(row, col + col_offset)
            tgt_cell.value = src_cell.value
            if src_cell.has_style:
                tgt_cell.font = copy(src_cell.font)
                tgt_cell.fill = copy(src_cell.fill)
                tgt_cell.border = copy(src_cell.border)
                tgt_cell.alignment = copy(src_cell.alignment)
                tgt_cell.number_format = src_cell.number_format
                tgt_cell.protection = copy(src_cell.protection)
            if src_cell.hyperlink:
                tgt_cell._hyperlink = copy(src_cell.hyperlink)
            if src_cell.comment:
                tgt_cell.comment = copy(src_cell.comment)

    for col in range(src_min_col, src_max_col + 1):
        src_letter = get_column_letter(col)
        tgt_letter = get_column_letter(col + col_offset)
        tgt_ws.column_dimensions[tgt_letter].width = src_ws.column_dimensions[src_letter].width
        tgt_ws.column_dimensions[tgt_letter].hidden = src_ws.column_dimensions[src_letter].hidden

    existing_merges = {str(item) for item in tgt_ws.merged_cells.ranges}
    for min_row, min_col, max_row, max_col in merge_targets:
        ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        if ref not in existing_merges:
            tgt_ws.merge_cells(ref)

def _resolve_block_repeat_target(ws, template_ws, mapping, base_row: int, base_col: int, index: int, copied_pages: set[int], defined_names: dict[str, dict], group_size: int = 1) -> tuple[int, int]:
    if getattr(mapping, "overflow_mode", None) != "sheet_right":
        return _resolve_repeat_target(ws, base_row, base_col, "down", index)

    page_settings = _resolve_page_settings(mapping, defined_names)
    block_start_row = page_settings["block_start_row"]
    block_end_row = page_settings["block_end_row"]
    block_start_col = page_settings["block_start_col"]
    block_end_col = page_settings["block_end_col"]
    page_start_col = page_settings["page_start_col"]
    page_end_col = page_settings["page_end_col"]
    if not all([block_start_row, block_end_row, block_start_col, block_end_col]):
        return _resolve_repeat_target(ws, base_row, base_col, "down", index)

    rows_per_block = max(block_end_row - block_start_row + 1, 1)
    start_col_idx = column_index_from_string(block_start_col)
    end_col_idx = column_index_from_string(block_end_col)
    body_width = max(end_col_idx - start_col_idx + 1, 1)
    page_start_col_idx = column_index_from_string(page_start_col) if page_start_col else start_col_idx
    page_end_col_idx = column_index_from_string(page_end_col) if page_end_col else end_col_idx
    page_width = max(page_end_col_idx - page_start_col_idx + 1, 1)
    subblock_count, subblock_width = _resolve_subblock_layout(mapping, body_width, page_width)
    block_fill_mode = group_size == 1 and body_width > 1

    if block_fill_mode:
        page_capacity = rows_per_block * body_width * subblock_count
        page_index = index // page_capacity
        page_item_index = index % page_capacity
        subblock_index = page_item_index // (rows_per_block * body_width)
        block_item_index = page_item_index % (rows_per_block * body_width)
        row_offset = block_item_index // body_width
        col_in_body = block_item_index % body_width
        col_offset = (page_index * page_width) + (subblock_index * subblock_width) + col_in_body
        target_col = start_col_idx + col_offset
    else:
        page_capacity = rows_per_block * subblock_count
        page_index = index // page_capacity
        page_item_index = index % page_capacity
        subblock_index = page_item_index // rows_per_block
        row_offset = page_item_index % rows_per_block
        col_offset = (page_index * page_width) + (subblock_index * subblock_width)
        target_col = base_col + col_offset

    if page_index > 0 and page_index not in copied_pages:
        _copy_cell_range(
            template_ws,
            ws,
            1,
            template_ws.max_row or block_end_row,
            page_start_col_idx,
            page_end_col_idx,
            col_offset=page_index * page_width,
        )
        copied_pages.add(page_index)

    target_row = block_start_row + row_offset
    while _is_merged_child(ws, target_row, target_col):
        target_row += 1
    return target_row, target_col


async def fetch_asset_data(asset_id: int, db: AsyncSession) -> dict:
    from app.models.master import EquipmentType, GroupNode, LocationNode, Person
    from sqlalchemy.orm import aliased

    Manager = aliased(Person)
    RepresentativeNic = aliased(AssetHwNic)

    stmt = (
        select(Asset, GroupNode, LocationNode, EquipmentType, Manager, RepresentativeNic)
        .outerjoin(GroupNode, Asset.group_id == GroupNode.id)
        .outerjoin(LocationNode, Asset.location_id == LocationNode.id)
        .outerjoin(EquipmentType, Asset.equipment_type_id == EquipmentType.id)
        .outerjoin(Manager, Asset.manager_id == Manager.id)
        .outerjoin(RepresentativeNic, Asset.representative_nic_id == RepresentativeNic.id)
        .where(Asset.id == asset_id)
    )
    row = (await db.execute(stmt)).first()
    if not row:
        return {}

    asset, group, location, equipment_type, manager, representative_nic = row
    data = {
        column.name: getattr(asset, column.name)
        for column in Asset.__table__.columns
        if column.name not in _EXCLUDED_FIELDS
    }
    data.update({
        "ip_address": representative_nic.ipv4_address if representative_nic else "",
        "group_name": group.name if group else "",
        "group_full_path": group.full_path if group else "",
        "location_name": location.name if location else "",
        "location_full_path": location.full_path if location else "",
        "equipment_type_name": equipment_type.name if equipment_type else "",
        "manager_name": manager.name if manager else "",
        "manager_title": manager.title if manager else "",
        "manager_contact": manager.contact if manager else "",
    })
    data.update(_flatten_custom_field_values(data.get("custom_fields_json")))
    return data


def _serialize_model_row(row: Any) -> dict:
    return {
        column.name: getattr(row, column.name)
        for column in row.__table__.columns
        if column.name not in _EXCLUDED_FIELDS
    }


async def _fetch_single_row(model: Any, asset_id: int, db: AsyncSession) -> dict:
    stmt = select(model).where(model.asset_id == asset_id).limit(1)
    row = (await db.execute(stmt)).scalar_one_or_none()
    if not row:
        return {}
    return _serialize_model_row(row)


async def _fetch_multi_rows(model: Any, asset_id: int, db: AsyncSession, order_by: Any | None = None) -> list[dict]:
    stmt = select(model).where(model.asset_id == asset_id)
    if order_by is not None:
        stmt = stmt.order_by(order_by)
    rows = (await db.execute(stmt)).scalars().all()
    return [_serialize_model_row(row) for row in rows]


async def fetch_data_source(data_source: str, asset_id: int, db: AsyncSession):
    if data_source == "asset":
        return await fetch_asset_data(asset_id, db)

    if data_source == "hw_system":
        return await _fetch_single_row(AssetHwSystem, asset_id, db)

    if data_source in _DATA_SOURCE_MODELS:
        model, is_repeatable, order_by = _DATA_SOURCE_MODELS[data_source]
        if is_repeatable:
            rows = await _fetch_multi_rows(model, asset_id, db, order_by)
            if data_source == "inspection":
                for row in rows:
                    if row.get("check_items") is not None:
                        row["check_items"] = str(row["check_items"])
            return rows
        return await _fetch_single_row(model, asset_id, db)

    if data_source == "today":
        now = datetime.now()
        weekdays = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        return {
            "date": now.strftime("%Y-%m-%d"),
            "year": str(now.year),
            "month": f"{now.month:02d}",
            "day": f"{now.day:02d}",
            "weekday": weekdays[now.weekday()],
        }

    if data_source == "static":
        return {}

    return {}


async def build_data_source_preview(
    data_source: str,
    asset_id: int,
    db: AsyncSession,
    max_rows: int = 5,
) -> dict[str, Any]:
    data = await fetch_data_source(data_source, asset_id, db)
    custom_field_keys = list((data or {}).get("custom_fields_json", {}).keys()) if isinstance(data, dict) else []
    catalog_items = [item for item in build_field_catalog(custom_field_keys) if item.data_source == data_source]
    field_names = [item.field for item in catalog_items]

    if isinstance(data, list):
        rows = []
        for index, item in enumerate(data[:max_rows], start=1):
            rows.append({
                "row_index": index,
                "values": {
                    field: _stringify_preview_value(item.get(field))
                    for field in field_names
                },
            })
        return {
            "asset_id": asset_id,
            "data_source": data_source,
            "is_repeatable": True,
            "total_rows": len(data),
            "truncated": len(data) > max_rows,
            "rows": rows,
        }

    row_values = {
        field: _stringify_preview_value(data.get(field) if isinstance(data, dict) else None)
        for field in field_names
    }
    return {
        "asset_id": asset_id,
        "data_source": data_source,
        "is_repeatable": False,
        "total_rows": 1 if any(value != "" for value in row_values.values()) else 0,
        "truncated": False,
        "rows": [{
            "row_index": 1,
            "values": row_values,
        }],
    }


def format_value(value, fmt) -> str:
    if value is None:
        return ""
    if isinstance(value, dict):
        return str(value)
    if fmt is None:
        return str(value)
    if "YYYY" in fmt and isinstance(value, (date, datetime)):
        return (
            fmt.replace("YYYY", str(value.year))
            .replace("MM", f"{value.month:02d}")
            .replace("DD", f"{value.day:02d}")
        )
    if "#" in fmt:
        try:
            return f"{float(value):,.0f}"
        except (ValueError, TypeError):
            return str(value)
    return str(value)


def extract_mapped_value(data: Any, field: str):
    if not field:
        return ""
    if isinstance(data, dict):
        if field in data:
            return data.get(field, "")
        if "." in field:
            current = data
            for part in field.split("."):
                if not isinstance(current, dict):
                    return ""
                current = current.get(part)
                if current is None:
                    return ""
            return current
        return data.get(field, "")
    if isinstance(data, list):
        if not data:
            return ""
        first_row = data[0]
        if isinstance(first_row, dict):
            if field in first_row:
                return first_row.get(field, "")
            return first_row.get(field, "")
    return ""


def _extract_list_field_values(data: Any, field: str) -> list[Any]:
    if not isinstance(data, list):
        return []
    values = []
    for item in data:
        if isinstance(item, dict):
            value = extract_mapped_value(item, field)
            if value not in (None, ""):
                values.append(value)
    return values


_ALLOWED_MATH_BINARY_OPERATORS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.FloorDiv: operator.floordiv,
    ast.Mod: operator.mod,
    ast.Pow: operator.pow,
}

_ALLOWED_MATH_UNARY_OPERATORS = {
    ast.UAdd: operator.pos,
    ast.USub: operator.neg,
}


def _coerce_numeric(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    return float(text)


def _eval_math_node(node, variables: dict[str, float]) -> float:
    if isinstance(node, ast.Expression):
        return _eval_math_node(node.body, variables)
    if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
        return float(node.value)
    if isinstance(node, ast.Name) and node.id in variables:
        return float(variables[node.id])
    if isinstance(node, ast.BinOp) and type(node.op) in _ALLOWED_MATH_BINARY_OPERATORS:
        left = _eval_math_node(node.left, variables)
        right = _eval_math_node(node.right, variables)
        return _ALLOWED_MATH_BINARY_OPERATORS[type(node.op)](left, right)
    if isinstance(node, ast.UnaryOp) and type(node.op) in _ALLOWED_MATH_UNARY_OPERATORS:
        return _ALLOWED_MATH_UNARY_OPERATORS[type(node.op)](_eval_math_node(node.operand, variables))
    raise ValueError("unsupported expression")


def _format_math_result(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.6f}".rstrip("0").rstrip(".")


def _render_output_template(output_template: str, variables: dict[str, Any], formatted_values: dict[str, str]) -> str:
    def replace_math(match: re.Match[str]) -> str:
        expr = match.group(1).strip()
        try:
            parsed = ast.parse(expr, mode="eval")
            numeric_variables = {key: _coerce_numeric(val) for key, val in variables.items()}
            result = _eval_math_node(parsed, numeric_variables)
            return _format_math_result(result)
        except Exception:
            return match.group(0)

    rendered = re.sub(r"\{\{\s*(.*?)\s*\}\}", replace_math, output_template)
    return (
        rendered
        .replace("{value}", formatted_values["value"])
        .replace("{secondary}", formatted_values["secondary"])
        .replace("{count}", formatted_values["count"])
    )


def resolve_mapping_value(data: Any, mapping, count_value: int | None = None) -> str:
    aggregate_mode = getattr(mapping, "aggregate_mode", None) or "value"
    raw_value = extract_mapped_value(data, mapping.field)
    secondary_raw_value = extract_mapped_value(data, getattr(mapping, "secondary_field", None))
    formatted_value = format_value(raw_value, mapping.format)
    secondary_formatted_value = format_value(secondary_raw_value, mapping.format)

    if isinstance(data, list):
        values = _extract_list_field_values(data, mapping.field)
        if aggregate_mode in {"value", "first"}:
            formatted_value = format_value(values[0] if values else "", mapping.format)
        elif aggregate_mode == "count":
            formatted_value = str(len(data))
        elif aggregate_mode == "join":
            formatted_value = ", ".join(format_value(value, mapping.format) for value in values)
        elif aggregate_mode == "join_unique":
            unique_values = list(dict.fromkeys(values))
            formatted_value = ", ".join(format_value(value, mapping.format) for value in unique_values)

    derived_count_value = 0
    if isinstance(data, list):
        derived_count_value = len(data)
    elif isinstance(data, dict):
        derived_count_value = 1 if raw_value not in (None, "") else 0
    if count_value is None:
        count_value = derived_count_value

    output_template = getattr(mapping, "output_template", None)
    if output_template:
        return _render_output_template(
            output_template,
            {
                "value": raw_value,
                "secondary": secondary_raw_value,
                "count": count_value,
            },
            {
                "value": formatted_value,
                "secondary": secondary_formatted_value,
                "count": str(count_value),
            },
        )
    return formatted_value


async def generate_form_report(template_id: int, asset_id: int, db: AsyncSession) -> tuple[str, str]:
    template = await db.get(ReportFormTemplate, template_id)
    if not template:
        raise ValueError("form template not found")

    stmt = (
        select(ReportFormMapping)
        .where(ReportFormMapping.template_id == template_id)
        .order_by(ReportFormMapping.sort_order)
    )
    mappings = (await db.execute(stmt)).scalars().all()

    data_cache: dict[str, Any] = {}

    async def get_data(source: str):
        if source not in data_cache:
            data_cache[source] = await fetch_data_source(source, asset_id, db)
        return data_cache[source]

    template_wb = load_workbook(template.file_path)
    wb = load_workbook(template.file_path)
    defined_names = _get_defined_name_map(template.file_path)

    for mapping in mappings:
        if mapping.repeat_direction:
            continue
        resolved_sheet_name, row, col = _resolve_mapping_anchor(mapping, defined_names)
        ws_name = resolved_sheet_name if resolved_sheet_name and resolved_sheet_name in wb.sheetnames else mapping.sheet_name
        ws = wb[ws_name] if ws_name and ws_name in wb.sheetnames else wb.active
        if mapping.data_source == "static":
            ws.cell(row, col).value = mapping.field
            continue
        data = await get_data(mapping.data_source)
        ws.cell(row, col).value = resolve_mapping_value(data, mapping)

    repeat_groups: dict[tuple[str, str, str, int], list[tuple[ReportFormMapping, int, int]]] = {}
    for mapping in mappings:
        if not mapping.repeat_direction:
            continue
        resolved_sheet_name, row, col = _resolve_mapping_anchor(mapping, defined_names)
        direction = mapping.repeat_direction or "down"
        anchor = row if direction == "down" else col
        key = (resolved_sheet_name or mapping.sheet_name or "", mapping.data_source, direction, anchor)
        repeat_groups.setdefault(key, []).append((mapping, row, col))

    for (_sheet_name, data_source, direction, _anchor), group in repeat_groups.items():
        data = await get_data(data_source)
        if not isinstance(data, list):
            continue
        max_rows = group[0][0].repeat_max_rows
        if max_rows is None or max_rows <= 0:
            max_rows = len(data)
        copied_pages: set[int] = set()
        for index, item in enumerate(data[:max_rows]):
            for mapping, base_row, base_col in group:
                ws_name = _sheet_name if _sheet_name and _sheet_name in wb.sheetnames else mapping.sheet_name
                ws = wb[ws_name] if ws_name and ws_name in wb.sheetnames else wb.active
                template_ws = template_wb[ws_name] if ws_name and ws_name in template_wb.sheetnames else template_wb.active
                if direction == "down":
                    target_row, target_col = _resolve_block_repeat_target(
                        ws,
                        template_ws,
                        mapping,
                        base_row,
                        base_col,
                        index,
                        copied_pages,
                        defined_names,
                        len(group),
                    )
                else:
                    target_row, target_col = _resolve_repeat_target(ws, base_row, base_col, direction, index)
                ws.cell(target_row, target_col).value = resolve_mapping_value(
                    item,
                    mapping,
                    count_value=len(data),
                )
                if index > 0:
                    src_cell = ws.cell(base_row, base_col)
                    tgt_cell = ws.cell(target_row, target_col)
                    if src_cell.has_style:
                        tgt_cell.font = copy(src_cell.font)
                        tgt_cell.fill = copy(src_cell.fill)
                        tgt_cell.border = copy(src_cell.border)
                        tgt_cell.alignment = copy(src_cell.alignment)
                        tgt_cell.number_format = src_cell.number_format

    asset_data = await get_data("asset")
    asset_code = asset_data.get("asset_code", "unknown") if isinstance(asset_data, dict) else "unknown"
    file_name = f"{template.name}_{asset_code}_{date.today()}.xlsx"

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return tmp.name, file_name
